import datetime
from openpyxl import load_workbook

number = "ИКБО-08-20"

today = datetime.datetime.today().weekday()

week = datetime.datetime.today().isocalendar()[1]-5

wb = load_workbook('../ИИТ_2 курс_21-22_весна_11нед.xlsx')

sheet = wb["ИИТ_2к_21-22_осень"]
for i in range(1,10):
    for j in range(1,10):
        print(sheet.cell(row = 2, column = 6).value)