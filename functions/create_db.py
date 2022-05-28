from work_with_sql import *
from openpyxl import load_workbook
import os
"""def cr_one(i):
    #os.remove(f'day_{i}.db')
    con = sql_connection()
    sql_table(con)
    ex = [0, ""]
    sql_insert(ex)"""
i = 0
def fill_day(num):
    name_of_heads = ["ИИИ", "ИИТ", "ИПТИП", "ИРЭИ", "ИТУ"]
    for head in name_of_heads:
        for cours in range(1, 6):

            if ((head == "ИИТ")and(cours == 4 or cours ==5))or((head == "ИПТИП")and(cours ==5))or((head == "ИТУ")and(cours == 4 or cours ==5)):
                continue

            wb = load_workbook(f'..\\sheds\\{head}_{cours} курс_21-22_весна_очка.xlsx')
            if (head == "ИИТ")and(cours == 1):
                sheet = wb["ИИТ_1к_21-22_осень"]
            elif (head == "ИИТ")and(cours == 2):
                sheet = wb["ИИТ_2к_21-22_осень"]
            elif (head == "ИИТ")and(cours == 3):
                sheet = wb["ИИТ_3к_21-22_весна"]
            else:
                sheet = wb["Лист1"]
            k = 0

            for j in range(6, sheet.max_column,5):
                print("max_column:", sheet.max_column)
                k+=1
                if (k%4 == 0):
                    continue
                print(j)
                group = sheet.cell(row = 2, column = j).value
                print(head)
                print(group)

                F = []

                for i in range(4+12*(num-1), 16+12*(num-1)):
                    if (sheet.cell(row=i, column=j).value) is None:
                        F+=[""]
                    else:
                        str0= ""
                        str1 = sheet.cell(row=i, column=j).value
                        if str1 is not None:
                            str0+=str1+";;"
                        else:
                            str0+=";;"
                        str2 = sheet.cell(row=i, column=j+1).value
                        if str2 is not None:
                            str0+=str2+";;"
                        else:
                            str0 += ";;"
                        str3 = str(sheet.cell(row=i, column=j+2).value)
                        if str3 is not None:
                            str0+=str3+";;"
                        else:
                            str0+=";;"
                        str4 = sheet.cell(row=i, column=j+3).value
                        if str4 is not None:
                            str0+=str4+";;"
                        else:
                            str0+=";;"
                        F += [str0]
                F.insert(0, group)
                #print(F)
                sql_insert_day(F, num)
def cr_mon(num):
    #os.remove("mondays.db")
    #con = sql_connection()

    create_table_of_day(num)
    #ex = ["", "","", "","", "","", "","", "","", "",""]
    #sql_insert_mon(ex)



def make_db_of_groups():
    os.remove("list_of_groups.db")
    create_list_of_groups()

    name_of_heads = ["ИИИ", "ИИТ", "ИПТИП", "ИРЭИ", "ИТУ"]
    for head in name_of_heads:
        for cours in range(1, 6):

            if ((head == "ИИТ") and (cours == 4 or cours == 5)) or (
                    (head == "ИПТИП") and (cours == 5)) or ((head == "ИТУ") and (cours == 4 or cours == 5)):
                continue

            wb = load_workbook(f'..\\sheds\\{head}_{cours} курс_21-22_весна_очка.xlsx')
            if (head == "ИИТ") and (cours == 1):
                sheet = wb["ИИТ_1к_21-22_осень"]
            elif (head == "ИИТ") and (cours == 2):
                sheet = wb["ИИТ_2к_21-22_осень"]
            elif (head == "ИИТ") and (cours == 3):
                sheet = wb["ИИТ_3к_21-22_весна"]
            else:
                sheet = wb["Лист1"]
            k = 0
            for j in range(6, sheet.max_column ,5):
                k+=1
                if (k%4 == 0):
                    continue
                global i
                i += 1
                group = sheet.cell(row = 2, column = j).value
                insert_in_list([i, group])
                print(i, group)
                print(sheet.max_column)

def make_table_of_groups_and_subs():
    os.remove("group_and_subs.db")
    create_table_group_and_subs()
    for num in range(1, 415):
        uniq_subs = ""
        group = find_in_table_list_of_groups(num)
        if (group == ""):
            continue
        for day in range(1,7):
            subs = find_subjects_by_group(group, day)
            for s in subs:
                s_s = s.split(";;")[0]
                if "…" in s_s:
                    continue
                if "Военная" in s_s:
                    s_s = "Военная подготовка"
                if "\n" in s_s:
                    s_d = s_s.split("\n")
                    for s_s in s_d:
                        if "н." in s_s:
                            s_s = s_s[s_s.find("н.")+3::]
                        if "(" in s_s:
                            s_s = s_s[:s_s.find(" ("):]
                        if (s_s != "")and (not(s_s in uniq_subs)):
                            uniq_subs += s_s +";;"
                else:
                    if "н." in s_s:
                        s_s = s_s[s_s.find("н.") + 3::]
                    if "(" in s_s:
                        s_s = s_s[:s_s.find(" ("):]
                    if (s_s != "") and (not (s_s in uniq_subs)):
                        uniq_subs += s_s + ";;"

        sql_insert_in_group_and_subs([group, uniq_subs])

"""make_db_of_groups()

for i in range(1,7):
    os.remove(f'..\\sheds\\db\\day_{i}.db')
    cr_mon(i)
    fill_day(i)
"""
make_table_of_groups_and_subs()